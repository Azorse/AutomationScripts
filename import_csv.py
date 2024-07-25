import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# Define file paths
csv_file_path = '/Users/danschissler/my_playwright_project/downloads/observation.csv'
excel_file_path = '/Users/danschissler/my_playwright_project/downloads/SIMworks_Data.xlsx'  
sheet_name = 'RawData'

# Read the .csv file
df_csv = pd.read_csv(csv_file_path)

#LCheck to see if exists and valid
if os.path.exists(excel_file_path):
    try:
        book = load_workbook(excel_file_path)
    except Exception as e:
        print(f"Error loading Excel file {excel_file_path}")
        exit(1)
else:
    print(f"Excel file does not exist at path: {excel_file_path}")
    exit(1)


#Check if sheet exists
if sheet_name in book.sheetnames:
    sheet = book[sheet_name]
else:
    sheet=book.create_sheet(sheet_name)

#find the last row
max_row = sheet.max_row

#write the csv data
for row in dataframe_to_rows(df_csv, index=False, header=(max_row == 1)):
    sheet.append(row)

#save file
book.save(excel_file_path)
book.close()

print(f"Contents of {csv_file_path} have been added to {excel_file_path} on the '{sheet_name}' sheet.")
