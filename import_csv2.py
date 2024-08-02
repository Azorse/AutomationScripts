import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# Load env variables
load_dotenv()

# Access env variables
email = os.environ.get('email')
password = os.environ.get('pwd')
csv_file_path = os.environ.get('csv')
excel_file_path = os.environ.get('excel')
sheet_name = 'RawData'

# Debugging statements
print(f"Email: {email}")
print(f"Password: {password}")
print(f"CSV File Path: {csv_file_path}")
print(f"Excel File Path: {excel_file_path}")

# Ensure the paths are correct
if not os.path.isfile(csv_file_path):
    raise FileNotFoundError(f"CSV file not found: {csv_file_path}")

if not os.path.isfile(excel_file_path):
    raise FileNotFoundError(f"Excel file not found: {excel_file_path}")

# Read the .csv file
df_csv = pd.read_csv(csv_file_path)

# Check if the Excel file exists and is valid
if os.path.exists(excel_file_path):
    try:
        book = load_workbook(excel_file_path)
    except Exception as e:
        print(f"Error loading Excel file {excel_file_path}: {e}")
        exit(1)
else:
    print(f"Excel file does not exist at path: {excel_file_path}")
    exit(1)

# Check if sheet exists
if sheet_name in book.sheetnames:
    sheet = book[sheet_name]
else:
    sheet = book.create_sheet(sheet_name)

# Find the first available empty row in a specific column (e.g., column A)
def get_first_empty_row(sheet, column):
    for row in range(1, sheet.max_row + 2):
        if sheet[f'{column}{row}'].value is None:
            return row

# Get the first available empty row in column A
first_empty_row = get_first_empty_row(sheet, 'A')

# Write the csv data starting from the first available empty row
for row_idx, row in enumerate(dataframe_to_rows(df_csv, index=False, header=(first_empty_row == 1)), start=first_empty_row):
    for col_idx, value in enumerate(row, start=1):
        sheet.cell(row=row_idx, column=col_idx, value=value)

# Columns to extend formulas
formula_columns = ['N', 'O', 'P', 'Q', 'R']

# Get the last row with data after adding new data
last_row_with_data = sheet.max_row

# Extend formulas in specified columns
for col in formula_columns:
    formula = sheet[f'{col}{first_empty_row - 1}'].value  # Get the formula from the last row before new data
    if formula and formula.startswith('='):  # Ensure it's a formula
        for row in range(first_empty_row, last_row_with_data + 1):
            sheet[f'{col}{row}'].value = formula

# Save file
book.save(excel_file_path)
book.close()

print(f"Contents of {csv_file_path} have been added to {excel_file_path} on the '{sheet_name}' sheet.")
